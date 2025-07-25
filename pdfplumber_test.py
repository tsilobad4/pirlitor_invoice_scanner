import pdfplumber
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

#normalizing the hyphens
def normalize_part_number(part):
    if not isinstance(part, str):
        return ""
    return part.replace("‐", "-").replace("–", "-").strip()

from copy import copy

# Use exactly the labels you show in the sheet:
MONTH_LABELS = ["JAN","FEB","MARCH","APRIL","MAY","JUNE","JUL","AUG","SEP","OCT","NOV","DEC"]
# A quick map from month‐number → label for lookups:
MONTH_MAP = {i+1: m for i, m in enumerate(MONTH_LABELS)}

# Stuff to make the row formatting work
def copy_row_format(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

part_entries = []           #  Empty list to hold extracted part data pulled from the PDF

with pdfplumber.open("test_invoice.pdf") as pdf:
    for i, page in enumerate(pdf.pages):
        # print(f"\n--- Page {i + 1} ---\n")

        #Breaking the block of text into  a list of lines splitting at every line break
        text = page.extract_text()
        lines = text.split('\n')
        #pattern = r"(?:Lot Price\s*)?\$\s?\d+\.\d{2}"  # Match "$ 5.50" or "Lot Price $ 150.00", just a pattern we will later be referencing
        pattern = r"(?:Lot Price\s*)?\$\s?[\d,]+\.\d{2}"


        # getting the invoice date (same for each page)
        invoice_date = None
        for line in lines:
            if line.startswith("Date:"):
                raw_date = line.replace("Date:", "").strip()
                parsed_date = datetime.strptime(raw_date, "%b/%d/%Y")  # e.g. Jun/07/2024
                invoice_date = parsed_date.strftime("%m.%d.%Y")        # → 06.07.2024
                break

        # print(f"Invoice Date: {invoice_date}")

        #Printing the lines that contain price info
        for line in lines: 
            
            if re.search(pattern, line):
                if re.search(pattern, line):
                    if any(keyword in line.lower() for keyword in ["subtotal", "hst", "total", "terms", "overdue"]):
                        continue  # Skip summary lines
               
                # print (line)
                tokens = line.split()
                #print (tokens)

                try:
                    amount = tokens[-2] + " " + tokens[-1]          # Last two tokens
                    unit_price = tokens[-4] + " " + tokens[-3]      # Handles dollar amount and lot price
                    quantity = tokens [-8]
                    part_number = " ".join(tokens[3:-8])              # Anything between token 3 and 8th last
                    
                    #print (part_number)

                    part_entries.append({
                        "date": invoice_date,
                        "part": part_number,
                        "qty": quantity,
                        "unit_price": unit_price,
                        "amount": amount

                    })

                except IndexError:
                    print("Skipped line due to unexpected format:", line)

#------------------------------------------------------------------

# Load in the workbook (Excel file) and find all part tables starting rows
wb = load_workbook("test_invoice_costing_UPDATED.xlsx")

# Select the specific worksheet we want to work with 
ws = wb["Sorting by part number"]

# Getting formatting templates from first part block
TEMPLATE_ROW_PART_HEADER = None
TEMPLATE_ROW_EMPTY = None
TEMPLATE_ROW_COLUMN_HEADERS = None
TEMPLATE_ROW_MONTH = None

for row in ws.iter_rows(min_row = 1, max_row = ws.max_row):
    for cell in row[:10]:
        if cell.value and isinstance(cell.value, str) and cell.value.strip().upper() == "PART NUMBER":
            TEMPLATE_ROW_PART_HEADER = cell.row
            TEMPLATE_ROW_EMPTY = cell.row + 1
            TEMPLATE_ROW_COLUMN_HEADERS = cell.row + 2
            TEMPLATE_ROW_MONTH = cell.row + 3               #Assumes JAN is the first month
            break

        if TEMPLATE_ROW_PART_HEADER:                         # Once we find first valid part table stop scanning further
            break


#--------------------------------------------------------------------

# Write entries to Excel
for entry in part_entries:
    part_number = normalize_part_number(entry["part"])
    # Search dynamically for the starting row of the part block
    start_row = None
    for row in ws.iter_rows(min_row = 1, max_row = ws.max_row):
        for cell in row[:10]:
            if cell.value and isinstance(cell.value, str) and cell.value.strip().upper() == "PART NUMBER":
                candidate = row[row.index(cell) + 1].value
                if normalize_part_number(candidate) == part_number:
                    start_row = cell.row
                    break

    #if the part is not already in the excel, we must add a new table
    if start_row is None:
        print(f"🆕 Creating new table for part {part_number}")
        new_table_start_row = ws.max_row + 3        # Leave two empty rows before new table

        # Copy PART NUMBER row
        ws.insert_rows(new_table_start_row)
        copy_row_format(ws, TEMPLATE_ROW_PART_HEADER, new_table_start_row)
        ws.cell(row = new_table_start_row, column = 2).value = "PART NUMBER"
        ws.cell(row = new_table_start_row, column = 3).value =normalize_part_number(entry["part"])
        new_table_start_row += 1

        # Copy empty spacer row
        ws.insert_rows(new_table_start_row)
        copy_row_format(ws, TEMPLATE_ROW_EMPTY, new_table_start_row)
        new_table_start_row += 1

        # Copy column header row
        ws.insert_rows(new_table_start_row)
        copy_row_format(ws, TEMPLATE_ROW_COLUMN_HEADERS, new_table_start_row)
        # Manually copy values for each header cell
        for col in range(1,8):
            header_val = ws.cell(row = TEMPLATE_ROW_COLUMN_HEADERS, column = col).value
            ws.cell(row = new_table_start_row, column = col).value = header_val
        new_table_start_row += 1

        # Copy and create rows for months JAN to DEC
        for month in ["JAN", "FEB", "MARCH", "APRIL", "MAY", "JUNE", "JUL", "AUG",
                      "SEP", "OCT", "NOV", "DEC"]:
            ws.insert_rows(new_table_start_row)
            copy_row_format(ws, TEMPLATE_ROW_MONTH, new_table_start_row)
            ws.cell(row=new_table_start_row, column=1).value = month
            new_table_start_row += 1

        # Update start_row: 1 (header) + 1 (spacer) + 1 (col headers) + 12 (months) = 15 rows
        start_row = new_table_start_row - 15

    target_row = None
    month_num = datetime.strptime(entry["date"], "%m.%d.%Y").month
    month_str = MONTH_MAP[month_num]
    
    current_row = start_row + 2
    insert_row = None               # where we will eventually right the data

    while current_row <= ws.max_row:


        col_a_val = ws.cell(row = current_row, column = 1).value
        col_b_val = ws.cell(row = current_row, column = 2).value

        #print(f"Row {current_row} | Col A: {col_a_val} | Col B: {col_b_val}")  # DEBUG LINE

        # Detect next part block, stop looking further
        if isinstance(col_b_val, str) and col_b_val.strip().upper() == "PART NUMBER":
            print("👋 Detected new part block at row", current_row)
            break

        # Check if we found the target month
        if isinstance(col_a_val, str) and col_a_val.strip().upper() == month_str:
            # If the cell right next to month is empty, use this row
            if not col_b_val:
                insert_row = current_row
                break
            else:
                # Scan downwards from this point until just before next month or part table
                search_row = current_row + 1
                while search_row <= ws.max_row:
                    next_col_a = ws.cell(row = search_row, column = 1).value
                    next_col_b = ws.cell(row = search_row, column = 2).value

                    # Stop if we hit a new month or next part table
                    if isinstance(next_col_a, str) and next_col_a.strip().upper() in MONTH_LABELS:
                        # Insert a blank row before this
                        ws.insert_rows(search_row)
                        copy_row_format(ws, search_row - 1, search_row)
                        insert_row = search_row
                        break
                    elif not next_col_b:             # If we find an empty row we can use
                        insert_row = search_row
                        break
                    search_row += 1
            break                                    # Stop looking after we found our month block


        current_row += 1

    # Now insert into Excel if we found a valid spot
    if insert_row:
        normalized_part = normalize_part_number(entry["part"])
        print(f"✅ Inserting at row {insert_row} for part '{normalized_part}' in month {month_str}")
        print(f"Inserting at row {insert_row}:")
        print(f"  Date: {entry['date']}")
        print(f"  Part: {normalize_part_number(entry['part'])}")
        print(f"  Qty: {entry['qty']}")
        print(f"  Unit Price: {entry['unit_price']}")
        print(f"  Amount: {entry['amount']}")

        ws.cell(row = insert_row, column = 2).value = entry["date"]
        ws.cell(row=insert_row, column=3).value = normalize_part_number(entry["part"])
        ws.cell(row = insert_row, column = 4).value = entry["qty"]

         # Simple insert if formatting was copied
        unit_price = entry["unit_price"].replace("$", "").strip()
        unit_price = unit_price.replace(",", "")
        amount = entry["amount"].replace("$", "").strip().replace(",", "")  # stripping the dollar sign and commas

        #ws.cell(row=insert_row, column=5).value = float(unit_price) if unit_price.lower() != "lot price" else "Lot Price"
        ws.cell(row=insert_row, column=5).value = float(unit_price.replace(",", "")) if unit_price.lower() != "lot price" else "Lot Price"
        ws.cell(row=insert_row, column=6).value = float(amount) if amount.lower() != "lot price" else "Lot Price"

    else: 
        print(f"Couldn't insert entry for part {part_number} and month {month_str}")


wb.save("test_invoice_costing_UPDATED.xlsx")
print("Excel updated and saved as 'test_invoice_costing_UPDATED.xlsx'")