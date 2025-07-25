from openpyxl import load_workbook
from copy import copy

MONTHS = ["JAN", "FEB", "MARCH", "APRIL", "MAY", "JUNE",
          "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]

def copy_row_format(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=source_row, column=col)
        tgt_cell = ws.cell(row=target_row, column=col)
        if src_cell.has_style:
            tgt_cell.font = copy(src_cell.font)
            tgt_cell.border = copy(src_cell.border)
            tgt_cell.fill = copy(src_cell.fill)
            tgt_cell.number_format = copy(src_cell.number_format)
            tgt_cell.protection = copy(src_cell.protection)
            tgt_cell.alignment = copy(src_cell.alignment)

def append_missing_months_openpyxl(filepath, max_tables=30):
    wb = load_workbook(filepath)
    ws = wb["Sorting by part number"]
    row = 1
    tables_done = 0

    while row <= ws.max_row and tables_done < max_tables:
        cell_val = ws.cell(row=row, column=2).value
        if isinstance(cell_val, str) and cell_val.strip().upper() == "PART NUMBER":
            part_number = ws.cell(row=row, column=3).value
            print(f"🔍 [{tables_done + 1}] Found part block: {part_number}")
            month_rows = {}
            month_start = row + 3

            # Scan next 20 rows to find existing months
            for i in range(month_start, month_start + 20):
                if i > ws.max_row:
                    break
                month_val = ws.cell(row=i, column=1).value
                if isinstance(month_val, str) and month_val.strip().upper() in MONTHS:
                    month_rows[month_val.strip().upper()] = i

            # Determine missing months
            missing_months = [m for m in MONTHS if m not in month_rows]
            if missing_months:
                print(f"📌 Missing months for {part_number}: {missing_months}")
                last_month_row = max(month_rows.values()) if month_rows else month_start - 1

                for m in missing_months:
                    insert_row = last_month_row + 1
                    ws.insert_rows(insert_row)
                    # Format the new row
                    copy_row_format(ws, last_month_row, insert_row)
                    ws.cell(row=insert_row, column=1).value = m
                    last_month_row += 1

                row = last_month_row + 1
            else:
                row += 15

            tables_done += 1
        else:
            row += 1

    output_file = "test_invoice_costing_UPDATED.xlsx"
    wb.save(output_file)
    print(f"✅ Done! Processed {tables_done} tables. Saved as '{output_file}'")

# Run it
append_missing_months_openpyxl("test_invoice_costing.xlsx", max_tables=30)
