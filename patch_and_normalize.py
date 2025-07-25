from openpyxl import load_workbook
from copy import copy

# 1) Load the “correct‑height but missing‑months” file
REF = "test_invoice_costing.xlsx"
wb = load_workbook(REF)
ws = wb["Sorting by part number"]

# 2) Sample header and normal heights from the first Data/Part + month row
for r in range(1, ws.max_row + 1):
    if ws.cell(row=r, column=2).value == "Data" and ws.cell(row=r, column=3).value == "Part":
        header_h = ws.row_dimensions[r].height or 15
        normal_h = ws.row_dimensions[r+1].height or 15
        break
else:
    raise RuntimeError("Couldn't find the Data/Part header in the reference file.")

# 3) Snapshot every PART NUMBER banner height so we can restore it later
banner_heights = {}
for r in range(1, ws.max_row + 1):
    if ws.cell(row=r, column=2).value == "PART NUMBER":
        banner_heights[r] = ws.row_dimensions[r].height or 15

# --- Month‑appending logic over the entire sheet ---
months = ["SEP", "OCT", "NOV", "DEC"]
labels = ["JAN", "FEB", "MARCH", "APRIL", "MAY", "JUNE", "JUL", "AUG"] + months
template_row = 5  # styling template (cells only; heights separate)

def is_month(v):
    return isinstance(v, str) and v.strip().upper() in labels

def copy_style(src, tgt):
    for c in range(1, ws.max_column + 1):
        s = ws.cell(row=src, column=c)
        t = ws.cell(row=tgt, column=c)
        t.font = copy(s.font)
        t.border = copy(s.border)
        t.fill = copy(s.fill)
        t.number_format = copy(s.number_format)
        t.protection = copy(s.protection)
        t.alignment = copy(s.alignment)

row = 1
while row <= ws.max_row:
    if ws.cell(row=row, column=2).value == "PART NUMBER":
        start = row + 3
        # find the last AUG label
        aug = None
        r = start
        while r <= ws.max_row:
            if ws.cell(row=r, column=2).value == "PART NUMBER":
                break
            if ws.cell(row=r, column=1).value == "AUG":
                aug = r
            r += 1
        if not aug:
            row += 1
            continue

        # find the end of that part block (last non-empty under AUG)
        end = aug
        r = aug + 1
        while r <= ws.max_row:
            if ws.cell(row=r, column=2).value == "PART NUMBER":
                break
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r, column=2).value
            if a or b:
                end = r
            else:
                break
            r += 1

        # collect existing months
        existing = {
            ws.cell(row=i, column=1).value.strip().upper()
            for i in range(start, end + 1)
            if is_month(ws.cell(row=i, column=1).value)
        }

        # insert missing months
        ins = end + 1
        for m in months:
            if m not in existing:
                ws.insert_rows(ins)
                copy_style(template_row, ins)
                ws.cell(row=ins, column=1).value = m
                ins += 1

        row = ins + 1
    else:
        row += 1

# --- Height‑normalization pass ---
def is_data_hdr(r):
    return (
        ws.cell(row=r, column=2).value == "Data"
        and ws.cell(row=r, column=3).value == "Part"
    )

for r in range(1, ws.max_row + 1):
    if r in banner_heights:
        # restore original banner height
        ws.row_dimensions[r].height = banner_heights[r]
    elif is_data_hdr(r):
        ws.row_dimensions[r].height = header_h
    else:
        ws.row_dimensions[r].height = normal_h

# 4) Save final result
OUT = "test_invoice_costing_FINAL.xlsx"
wb.save(OUT)
print(f"✅ All parts processed. Output written to '{OUT}'.")
